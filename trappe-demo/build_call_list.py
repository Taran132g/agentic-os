#!/usr/bin/env python3
"""Piontrix call list — 15 local independents, fit-scored & color-coded.

Screening rubric (applied here + on every future batch):
  SKIP if they already run a tool that does my job — Toast/Square ordering,
  review/messaging suites (Podium/Birdeye/Weave/Solutionreach), full
  booking+marketing platforms (Mindbody/Boulevard/Vagaro/fitDEGREE/OpenTable
  w/ reminders), chains/DSOs, or no reachable owner.
  PRIORITIZE visible GAPS that my workflows fill — good rating + few reviews,
  call-to-book (no widget), "couldn't reach them" reviews, no review replies,
  basic/FB-only site, owner-run single location, appointment-driven, newer.
Fit: A = clean gaps, no overlap · B = partial overlap (lead w/ uncovered
workflow) · C = skip/deprioritize (overlapping tool or unreachable).
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = Path.home() / "Downloads" / "Piontrix_Call_List_2026-06-16.xlsx"

# color: red = call back · purple = to call · gray = skip/deprioritize (filtered out)
ROWS = [
    {"order": 1, "color": "purple", "fit": "B – partial overlap", "company": "TruFit Pilates", "vertical": "Boutique fitness studio",
     "phone": "(484) 924-9931", "website": "trufitpilates.com", "address": "235 Bridge St, Phoenixville",
     "rating": "Verify on Google (newer)", "window": "11:30am–1:00pm",
     "flags": "Runs FRESHA (booking + reminders already covered). Lead Reactivation + Missed-call, not reminders.",
     "tech": "Fresha (online booking + payments). Reformer/strength classes.",
     "workflows": "Reactivation (churn) · Class/late-cancel recovery · Missed-call · Reputation",
     "hook": "\"Boutique studios live on whether members keep booking. The ones who drift off for two weeks usually cancel — I catch them and pull them back.\"",
     "pain": "Class churn; empty reformer spots; thin reviews.", "roi": "1 retained member ≈ $180–220/mo.",
     "who": "Owner/lead instructor — midday between classes."},
    {"order": 2, "color": "purple", "fit": "A – strong", "company": "The Mayor Salon", "vertical": "Hair salon (boutique)",
     "phone": "Call (484) 854-6954 · Text (484) 902-8263", "website": "themayorsalon.com", "address": "430 W Main St, Collegeville",
     "rating": "Verify on Google (newer)", "window": "1:00–2:00pm",
     "flags": "No overlapping suite. TEXTS clients MANUALLY = perfect for owner-send automation. Newer = review gap.",
     "tech": "Appointment-based; texts clients manually.",
     "workflows": "Reminders (no-show recovery) · Reactivation (never-rebooked) · Reputation · Missed-call",
     "hook": "\"A no-show is an hour you can't sell back. And the clients who loved you once but never rebooked — I bring them back with a text from your own number.\"",
     "pain": "No-shows; rebooking; building reviews as a newer salon.", "roi": "Recovered no-show ≈ $60–120.",
     "who": "Owner — they already text clients."},
    {"order": 3, "color": "red", "fit": "A – strong (VERIFY)", "company": "Advanced Dentistry of Phoenixville", "vertical": "Dental practice",
     "phone": "(610) 935-7509", "website": "phoenixvilledentistry.com", "address": "119 Nutt Rd, Phoenixville",
     "rating": "~4.9 but LOW volume (~11 Yelp)", "window": "1:30–3:00pm (Fri closes 3)",
     "flags": "Huge review gap. ⚠ VERIFY they don't already run Weave / Solutionreach / RevenueWell (dental suites cover reminders+recall+reviews). Also confirm owner-operated vs DSO.",
     "tech": "Dental PMS (booking). Website = contact form only.",
     "workflows": "Recall reactivation (biggest $) · Reminders · Reputation (4.9 but few reviews) · Missed-call/Voice",
     "hook": "\"You're at 4.9 stars but a handful of reviews — rivals have hundreds. And every overdue recall is a cleaning you're not billing. I bring both back automatically.\"",
     "pain": "No-shows; overdue recalls; thin reviews.", "roi": "~23 recalls ≈ $4,600; no-show ≈ $200+.",
     "who": "Office manager — confirm owner vs DSO early."},
    {"order": 5, "color": "purple", "fit": "A – strong", "company": "Trappe Tavern", "vertical": "Bar / restaurant (tavern)",
     "phone": "(610) 489-8686", "website": "trappetavern.com", "address": "416 W Main St, Trappe",
     "rating": "4.6 · 693 reviews", "window": "2:30–4:30pm (best window)",
     "flags": "No booking, no ordering, no comms suite. Clean missed-call + reactivation gaps. Reputation already strong — don't lead there.",
     "tech": "No online ordering, no booking. Reservations by phone. (Your sharpest demo.)",
     "workflows": "Missed-call/Voice (no host after 9pm) · Reactivation (regulars from res log) · Digest (estimated)",
     "hook": "\"On a packed Friday after 9, who's grabbing the phone? Every missed call is a party booking somewhere else. I catch them — or have the AI answer live.\"",
     "pain": "Missed after-hours/rush calls; regulars drifting; no res system.", "roi": "1 large party ≈ $200–500.",
     "who": "Owner — best 2:30–4:30."},
    {"order": 6, "color": "purple", "fit": "B – partial overlap", "company": "Bistro on Bridge", "vertical": "Gastropub / restaurant",
     "phone": "(610) 935-7141", "website": "bistroonbridge.com", "address": "212 Bridge St, Phoenixville",
     "rating": "4.4 · 1,776 reviews (strong)", "window": "2:00–4:30pm",
     "flags": "Reviews strong (NOT a review gap) + likely OpenTable (reminders covered). Lead Missed-call + private-event capture.",
     "tech": "Reservations + rooftop garden + private events. Since 2009.",
     "workflows": "Missed-call/Voice · Private-event / large-party capture · Reactivation",
     "hook": "\"When the bar's slammed, who's catching the reservation and private-event calls? I make sure none ring out.\"",
     "pain": "Missed calls during rush; event inquiries sitting.", "roi": "1 booked private event = $$$.",
     "who": "Owner/GM — mid-afternoon."},
    {"order": 7, "color": "purple", "fit": "B – partial overlap", "company": "Black Lab Bistro", "vertical": "Fine dining (BYOB)",
     "phone": "(610) 935-5988", "website": "blacklabbistro.net", "address": "248 Bridge St, Phoenixville",
     "rating": "Verify (strong on OpenTable)", "window": "2:30–4:30pm",
     "flags": "OpenTable likely covers reservation reminders. Lead Missed-call + Reactivation.",
     "tech": "Reservation-driven (OpenTable). BYOB.",
     "workflows": "Reminders (no-show recovery) · Reactivation · Missed-call",
     "hook": "\"Fine dining lives on the reservation book — a Saturday no-show is a real hit. I confirm covers and bring back guests who haven't been in.\"",
     "pain": "Reservation no-shows; missed booking calls; lapsed regulars.", "roi": "Saved 2-top ≈ $100–200.",
     "who": "Owner/chef-owner — pre-dinner."},
    {"order": 8, "color": "purple", "fit": "A – strong", "company": "Bluebird Distilling", "vertical": "Craft distillery / cocktail bar",
     "phone": "(610) 933-7827", "website": "bluebirddistilling.com", "address": "100 Bridge St, Phoenixville",
     "rating": "Verify on Google", "window": "1:00–3:30pm (before 4pm open)",
     "flags": "Event/tasting bookings likely manual = real gap. Verify any booking tool. Opens 4pm → calls before that go nowhere.",
     "tech": "Tasting room + private events + bottle sales.",
     "workflows": "Event/tasting/private-party booking · Missed-call/Voice · Reactivation · Reputation",
     "hook": "\"Tastings and private events are your margin — every missed booking call is lost. I answer them and capture the event leads.\"",
     "pain": "Missed event calls; event capture.", "roi": "1 private tasting/event = hundreds.",
     "who": "Owner/events lead — early afternoon before open."},
    {"order": 9, "color": "purple", "fit": "B – partial overlap", "company": "Sweet Brew Cafe", "vertical": "Café / bagels",
     "phone": "(484) 927-4795", "website": "sweetbrewpxv.com", "address": "158 Bridge St, Phoenixville",
     "rating": "Verify (~26 Yelp)", "window": "10:30am–12:30pm (closes 1pm)",
     "flags": "On SQUARE (ordering/loyalty partial — lighter than Toast). Lead catering Missed-call + Reviews.",
     "tech": "Square (first-party ordering) + DoorDash/UberEats.",
     "workflows": "Reputation (reviews) · Reactivation (loyalty) · Missed-call (catering/large orders)",
     "hook": "\"Morning rush is brutal — catering and large-order calls that ring out are pure lost revenue. Plus I keep your regulars coming back.\"",
     "pain": "Unanswered calls at rush; catering capture; loyalty.", "roi": "Catering + repeat regulars.",
     "who": "Owner — after morning rush, before 1pm."},
    {"order": 10, "color": "purple", "fit": "B – verify", "company": "Gwendolynn's Salon & Spa", "vertical": "Full-service salon & spa",
     "phone": "(610) 495-9300", "website": "gwendolynnssalonspa.com", "address": "628 W Ridge Pike, Limerick",
     "rating": "Verify on Google", "window": "1:30–3:30pm (Sat till 2)",
     "flags": "⚠ VERIFY booking system (Vagaro/Boulevard cover reminders). If basic, gaps in reactivation/reviews/missed-call.",
     "tech": "Appointment-based salon/spa. Mon–Fri 9–9.",
     "workflows": "Reminders (no-show recovery) · Reactivation · Reputation · Missed-call",
     "hook": "\"A no-show is an hour you can't sell, and the clients who drift off cost the most — I bring them back with a text from your own number.\"",
     "pain": "No-shows; rebooking; reviews.", "roi": "No-show ≈ $50–150; reactivated regular recurring.",
     "who": "Owner/front desk."},
    {"order": 11, "color": "purple", "fit": "B – partial overlap", "company": "Providence Spa & Nails", "vertical": "Nail salon / spa",
     "phone": "(610) 409-6300", "website": "(Yelp/Fresha listing)", "address": "500 Broad St, Collegeville",
     "rating": "~115 Yelp reviews", "window": "1:00–3:00pm",
     "flags": "On FRESHA (booking + reminders covered). Lead Reactivation (rebooking cadence) + Missed-call.",
     "tech": "Walk-ins + appointments (Fresha). High-frequency rebooking.",
     "workflows": "Reactivation (rebooking) · Reminders · Missed-call · Reputation",
     "hook": "\"Nail clients rebook every few weeks — when they lapse, a quick text brings them back. And on busy days I catch the booking calls you miss.\"",
     "pain": "Lapsed rebookers; missed calls on busy days.", "roi": "Reactivated regular = recurring visits.",
     "who": "Owner/manager — midday."},
    {"order": 12, "color": "purple", "fit": "B – verify", "company": "CoreFit Training Studio", "vertical": "Boutique fitness studio",
     "phone": "(610) 213-5110", "website": "corefittraining.net", "address": "332 Myrtle St, Royersford",
     "rating": "Verify on Google", "window": "11:00am–1:00pm / 1:30–3:30pm",
     "flags": "⚠ Verify class platform (MindBody/fitDEGREE cover booking+reminders). Lead Reactivation (churn). Owner-run.",
     "tech": "Class-based (kettlebell/TRX/hot yoga). Founder Liz Gilinger.",
     "workflows": "Reactivation (class churn) · Reminders · Reputation",
     "hook": "\"Small studios live on retention — I catch the members who stop showing before they cancel and pull them back.\"",
     "pain": "Class churn; fill; reviews.", "roi": "1 retained member ≈ $100–180/mo.",
     "who": "Owner (Liz) — between class blocks."},
    {"order": 13, "color": "purple", "fit": "B – hard to reach", "company": "Main Street Fitness", "vertical": "Boutique gym",
     "phone": "No public # — msnfitness.co (msnfitnessinfo@gmail.com)", "website": "msnfitness.co",
     "address": "360 E Main St, Collegeville", "rating": "Verify on Google", "window": "Email/DM first",
     "flags": "Has online booking; NO public phone (harder to reach). Intro via email/DM, then ask for a call. Retention gap.",
     "tech": "Boutique gym — WOD/small-group/PT. Booking at msnfitness.co/booking.",
     "workflows": "Reactivation (churn) · Class reminders · Reputation",
     "hook": "\"The members who quietly stop showing before they cancel — I catch them and pull them back.\"",
     "pain": "Churn; class fill; reviews.", "roi": "Retention = recurring $.",
     "who": "Owner — email/DM first (no listed phone)."},
    {"order": 15, "color": "purple", "fit": "A – gaps (VERIFY reachable)", "company": "Swedeworks Automotive", "vertical": "Auto repair [from your outreach]",
     "phone": "Verify on Google (Phoenixville)", "website": "(verify)", "address": "Phoenixville",
     "rating": "Verify on Google", "window": "8:00–10:00am or midday",
     "flags": "Minimal online presence = big missed-call/reviews/service-reminder gaps. ⚠ But verify they're reachable & serious (only a Yellow Pages listing found). NOT a salon — it's auto repair.",
     "tech": "Independent auto-repair shop.",
     "workflows": "Missed-call/Voice (missed call = lost job) · Service-due reminders (reactivation) · Reputation",
     "hook": "\"Every call you miss while under a car is a repair going to the shop down the road — I catch them, and bring customers back when they're due for service.\"",
     "pain": "Missed calls; no service follow-up; reviews.", "roi": "1 recovered job = $100s; reminders = repeat.",
     "who": "Owner — look up the number first."},

    # ── Batch 3 (5 more, filter-passed local independents) ──
    {"order": 16, "color": "purple", "fit": "B – partial overlap", "company": "Molly Maguire's Irish Restaurant & Pub", "vertical": "Irish pub / restaurant",
     "phone": "(610) 933-9550", "website": "mollymaguiresphoenixville.com", "address": "197 Bridge St, Phoenixville",
     "rating": "~172 Tripadvisor reviews (verify Google)", "window": "2:00–4:30pm",
     "flags": "Live music + private events + likely OpenTable. Lead Missed-call + private-event/live-music booking capture.",
     "tech": "Reservations + private events + live entertainment. Downtown corner spot.",
     "workflows": "Missed-call/Voice · Private-event & party-booking capture · Reactivation",
     "hook": "\"You've got live music and private events — every booking call that rings out during service is money gone. I make sure they're all answered.\"",
     "pain": "Missed calls during service; event/party inquiries; busy downtown corner.", "roi": "1 private event/party = $$$.",
     "who": "Owner/GM — mid-afternoon."},
    {"order": 17, "color": "purple", "fit": "A – strong", "company": "Spinal Care Chiropractic", "vertical": "Chiropractor",
     "phone": "(610) 489-8800", "website": "(Yelp/listing)", "address": "109 2nd Ave, Collegeville",
     "rating": "~20 Yelp reviews (LOW = gap)", "window": "1:30–3:30pm (between adjustments)",
     "flags": "Low review count = reputation gap. Recall/reactivation (patients who fall off care plans) + reminders + missed-call. Verify any chiro EHR recall (ChiroTouch) — reviews + reactivation still open.",
     "tech": "Chiropractic practice; appointment-based.",
     "workflows": "Reactivation (lapsed care plans) · Reminders (no-show recovery) · Reputation (few reviews) · Missed-call",
     "hook": "\"Patients who fall off their care plan are your biggest leak — and you've got a handful of reviews while you should have hundreds. I bring both back automatically.\"",
     "pain": "Patients dropping care plans; no-shows; thin reviews.", "roi": "Reactivated care plan = multiple visits.",
     "who": "Owner-chiropractor / front desk."},
    {"order": 18, "color": "purple", "fit": "B – verify", "company": "Pure Bliss Medical Spa", "vertical": "Med spa / aesthetics",
     "phone": "(484) 919-3058", "website": "pureblissmed-spa.com", "address": "507 Kimberton Rd, Phoenixville",
     "rating": "Verify on Google", "window": "1:00–3:00pm",
     "flags": "High-value (Botox/fillers). ⚠ Verify booking tool (Vagaro/GlossGenius cover reminders). Reactivation on treatment cycles (Botox ~every 3–4 mo) is the money.",
     "tech": "Aesthetics / injectables / wellness; appointment-based.",
     "workflows": "Reactivation (treatment-cycle recall) · Reminders · Reputation · Missed-call",
     "hook": "\"Your Botox clients are due again every few months — when they don't rebook, that's $300+ walking out. I bring them back right on cycle.\"",
     "pain": "Clients lapsing between cycles; missed booking calls; reviews.", "roi": "1 reactivated injectable client = $300–600/cycle.",
     "who": "Owner/PA — early afternoon."},
    {"order": 19, "color": "purple", "fit": "B – partial overlap", "company": "Hickory Springs Farm (Pet Grooming)", "vertical": "Pet grooming / boarding",
     "phone": "(610) 933-9584", "website": "hickoryspringsfarm.com", "address": "Phoenixville",
     "rating": "Verify on Google", "window": "9:00–11:00am or midday",
     "flags": "Family-owned since 2007. Grooming clients rebook every 6–8 wks → reactivation + reminders fit. Missed booking calls during grooming.",
     "tech": "Pet grooming + boarding.",
     "workflows": "Reactivation (rebooking cadence) · Reminders (appointment confirms) · Missed-call · Reputation",
     "hook": "\"Grooming clients should come back every 6–8 weeks — when they lapse, a quick text brings them in. And the calls you miss mid-groom are bookings gone.\"",
     "pain": "Lapsed grooming rebookers; missed calls while grooming; reviews.", "roi": "Reactivated groom client = recurring every 6–8 wks.",
     "who": "Owner — call morning before appointments ramp."},
    {"order": 20, "color": "purple", "fit": "B – partial overlap", "company": "Eddie The Barber", "vertical": "Barbershop",
     "phone": "(484) 630-7506", "website": "eddiethebarber.com", "address": "119 Main St, Ste 4, Phoenixville",
     "rating": "Verify on Google", "window": "Midday (closed Mon; Tue–Wed close 2pm)",
     "flags": "Likely on a booking app (Booksy/Squire → reminders covered). Lead Reactivation (lapsed regulars) + Reviews + Missed-call.",
     "tech": "Barbershop; appointment-based (likely Booksy/Squire).",
     "workflows": "Reactivation (lapsed regulars) · Reputation · Missed-call",
     "hook": "\"The regulars who came every few weeks and drifted off — I bring them back with a text, and catch the booking calls you miss mid-cut.\"",
     "pain": "Lapsed regulars; missed calls mid-cut; reviews.", "roi": "Reactivated regular = recurring cuts.",
     "who": "Owner (Eddie) — call midday; closed Mondays."},
]

wb = openpyxl.Workbook()
HEAD_FILL = PatternFill("solid", fgColor="2F2A24")
HEAD_FONT = Font(bold=True, color="E9B873", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))
FILLS = {"red": PatternFill("solid", fgColor="F4CCCC"),
         "purple": PatternFill("solid", fgColor="E4D5F7"),
         "gray": PatternFill("solid", fgColor="D9D9D9")}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 22


def paint(ws, ncols, row_height=None):
    for ri, r in enumerate(ROWS, start=2):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=ri, column=c)
            cell.alignment = WRAP; cell.border = THIN; cell.fill = FILLS[r["color"]]
        if row_height:
            ws.row_dimensions[ri].height = row_height


# ── Sheet 1: Call List ──
ws1 = wb.active; ws1.title = "Call List"
cols1 = [("#", 4), ("Company", 32), ("Vertical", 22), ("Phone", 30), ("Website", 24),
         ("Address", 28), ("Rating / reviews", 22), ("Call window", 24),
         ("Fit", 22), ("Already runs / verify", 50), ("Called?", 9), ("Outcome / next step", 26)]
ws1.append([c[0] for c in cols1])
for idx, r in enumerate(ROWS, 1):
    ws1.append([idx, r["company"], r["vertical"], r["phone"], r["website"], r["address"],
                r["rating"], r["window"], r["fit"], r["flags"], "", ""])
for i, (_, w) in enumerate(cols1, 1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
paint(ws1, len(cols1), row_height=70); style_header(ws1, len(cols1))

# ── Sheet 2: Analysis & Pitch ──
ws2 = wb.create_sheet("Analysis & Pitch")
cols2 = [("#", 4), ("Company", 28), ("Tech they already run", 36), ("Top workflows to pitch", 42),
         ("The hook (opener)", 48), ("Pain points", 34), ("ROI angle", 30), ("Who to ask for", 28)]
ws2.append([c[0] for c in cols2])
for idx, r in enumerate(ROWS, 1):
    ws2.append([idx, r["company"], r["tech"], r["workflows"], r["hook"], r["pain"], r["roi"], r["who"]])
for i, (_, w) in enumerate(cols2, 1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
paint(ws2, len(cols2), row_height=92); style_header(ws2, len(cols2))

# ── Sheet 3: Legend + screening rubric ──
ws3 = wb.create_sheet("Legend & Filter")
ws3.column_dimensions["A"].width = 90
L = [
    ("Color key", True), ("", False),
    ("  Red = call back / follow-up", "red"),
    ("  Purple = to call (qualified)", "purple"),
    ("  Gray = skip / deprioritize (filtered out — already served or unreachable)", "gray"),
    ("", False),
    ("Fit:  A = clean gaps, no overlap  ·  B = partial overlap (lead w/ uncovered workflow)  ·  C = skip", False),
    ("", False),
    ("SCREENING FILTER (applied here + on every future batch)", True),
    ("", False),
    ("SKIP if they already run a tool that does my job:", True),
    ("  • Toast / Square online ordering (bundled loyalty/marketing)", False),
    ("  • Review/messaging suites: Podium, Birdeye, Weave, Solutionreach, Demandforce", False),
    ("  • Booking+marketing platforms: Mindbody, Boulevard, Vagaro, fitDEGREE, Jane, OpenTable/Resy w/ reminders", False),
    ("  • Chains / franchises / DSO-owned (no local decision-maker)", False),
    ("  • Big enough for in-house marketing/IT; or no reachable owner / dead online presence", False),
    ("", False),
    ("PRIORITIZE visible gaps my workflows fill:", True),
    ("  • Good rating but LOW review count (e.g. 4.5+, <50)", False),
    ("  • 'Call to book' / reservations by phone, no booking widget", False),
    ("  • Reviews complaining 'couldn't get through / no one answered'", False),
    ("  • Doesn't reply to Google reviews; slow review velocity", False),
    ("  • Basic / Facebook-only / old website, no chat or booking", False),
    ("  • Owner-operated single location; appointment/reservation-driven; newer (<2 yrs)", False),
    ("  • Third-party-delivery-only; active-but-manual (posts/texts by hand)", False),
    ("", False),
    ("RULE OF THUMB: skip if a tool already does my job; prioritize if I can SEE the leak my workflows fill.", True),
]
for ri, item in enumerate(L, start=1):
    text, style = (item[0], item[1])
    c = ws3.cell(row=ri, column=1, value=text)
    if style is True:
        c.font = Font(bold=True, size=12 if ri == 1 else 11)
    elif style in FILLS:
        c.fill = FILLS[style]

OUT.parent.mkdir(exist_ok=True)
if OUT.exists():
    raise SystemExit(
        f"Refusing to overwrite {OUT.name} — it's the curated source of truth now.\n"
        "Edit the spreadsheet in place (add/remove specific rows); don't regenerate.")
wb.save(OUT)
print("saved:", OUT, "| rows:", len(ROWS),
      "| skip(gray):", sum(1 for r in ROWS if r["color"] == "gray"))
